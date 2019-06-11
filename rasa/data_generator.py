import json
from openpyxl import load_workbook
from os.path import join as path_join


rasa_data = {
    "rasa_nlu_data": {
        "common_examples": [],
        "regex_features": [],
        "lookup_tables": [],
        "entity_synonyms": []
    }
}

EXCEL_DIR = "excel"
file_to_intent = {
    "inheritance.xlsx": "know_inherited_diseases",
    "origin.xlsx": "know_ancestry",
}
filepath_to_intent = {}

# Adding EXCEL_DIR to each filename
for key in file_to_intent.keys():
    new_key = path_join(".", EXCEL_DIR, key)
    filepath_to_intent[new_key] = file_to_intent[key]

intent_to_text = {}


def merge_lists_from_dict_except_key(my_dict, excepted_key):
    """
    Merges lists that are in dictionary (except one) into one dict
    :param dict my_dict: dictionary key -> list
    :param string excepted_key: the key of list we shouldn't include in sum list
    :return list: sum list
    """
    res = []
    for key, value in my_dict.items():
        if key != excepted_key:
            res += value
    return res


def create_data():
    """
    Converts user searches data from csv files into rasa chatbot training data and saves it
    """
    intent_to_text_new = {}

    # Extracting text searches from csv files for each intent
    for file_name, intent in filepath_to_intent.items():
        intent_to_text[intent] = []
        wb = load_workbook(filename=file_name)
        def_sheet = wb.active
        for cell in def_sheet['B']:
            text = cell.value
            intent_to_text[intent].append(text)

    sets = []

    # Removing text searches appearing in several csv files (intents) like 'dna', 'genotek' for more productivity
    for intent, text_list in intent_to_text.items():
        other_words = merge_lists_from_dict_except_key(intent_to_text, intent)
        not_allowed = list(set(text_list).intersection(other_words))
        print(intent, not_allowed)
        sets.append(not_allowed)
        allowed_words = list(set(text_list).symmetric_difference(other_words))
        intent_to_text_new[intent] = allowed_words

    # Converting text searches into rasa nlu training data
    for file_name, intent in filepath_to_intent.items():
        wb = load_workbook(filename=file_name)
        def_sheet = wb.active
        for cell in def_sheet['B']:
            text = cell.value
            if text in intent_to_text_new[intent]:
                example = {
                    "text": text,
                    "intent": intent,
                    "entities": []
                }
                rasa_data["rasa_nlu_data"]["common_examples"].append(example)

    # Saving training data
    with open("data/data.json", "w") as f:
        json.dump(rasa_data, f, indent=4, ensure_ascii=False)


if __name__ == '__main__':
    create_data()
